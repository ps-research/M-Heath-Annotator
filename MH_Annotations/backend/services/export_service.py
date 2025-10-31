"""
Export generation service.
"""

import json
import tempfile
from pathlib import Path
from typing import Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from backend.services.data_service import DataService


class ExportService:
    """Service for generating exports."""

    def __init__(self):
        """Initialize export service."""
        self.data_service = DataService()

    def generate_export(self, request: Dict[str, Any]) -> str:
        """
        Generate export file.
        
        Args:
            request: Export request dictionary
            
        Returns:
            Path to generated export file
        """
        export_format = request.get("format", "json")
        
        if export_format == "json":
            return self.generate_json_export(request)
        elif export_format == "excel":
            return self.generate_excel_export(request)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

    def generate_json_export(self, request: Dict[str, Any]) -> str:
        """Generate JSON export."""
        filters = request.get("filters", {})
        
        # Get all data (no pagination)
        filters["page_size"] = 100000
        result = self.data_service.get_annotations(filters)
        
        # Create temp file
        temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
        
        # Write JSON
        json.dump(result["items"], temp_file, indent=2, ensure_ascii=False)
        temp_file.close()
        
        return temp_file.name

    def generate_excel_export(self, request: Dict[str, Any]) -> str:
        """Generate Excel export."""
        filters = request.get("filters", {})
        excel_options = request.get("excel_options", {})
        multi_sheet = excel_options.get("multi_sheet", True)
        include_summary = excel_options.get("include_summary", True)
        
        # Get all data
        filters["page_size"] = 100000
        result = self.data_service.get_annotations(filters)
        items = result["items"]
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        if multi_sheet:
            # Create sheet per domain
            domains = set(item["domain"] for item in items)
            
            for domain in sorted(domains):
                domain_items = [item for item in items if item["domain"] == domain]
                self._create_domain_sheet(wb, domain, domain_items)
        else:
            # Single sheet with all data
            ws = wb.create_sheet("All Annotations")
            self._write_items_to_sheet(ws, items)
        
        # Add summary sheet
        if include_summary:
            self._create_summary_sheet(wb, items)
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name

    def _create_domain_sheet(self, wb: Workbook, domain: str, items: list):
        """Create a sheet for a specific domain."""
        ws = wb.create_sheet(domain.capitalize())
        self._write_items_to_sheet(ws, items)

    def _write_items_to_sheet(self, ws, items: list):
        """Write items to a worksheet."""
        # Headers
        headers = ["ID", "Text", "Label", "Malformed", "Annotator", "Timestamp"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for item in items:
            ws.append([
                item.get("id", ""),
                item.get("text", "")[:200],  # Truncate long text
                item.get("label", ""),
                "Yes" if item.get("malformed", False) else "No",
                item.get("annotator_id", ""),
                item.get("timestamp", "")
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20

    def _create_summary_sheet(self, wb: Workbook, items: list):
        """Create summary statistics sheet."""
        ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Title
        ws.append(["Annotation Export Summary"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        
        # Basic stats
        total = len(items)
        malformed = sum(1 for item in items if item.get("malformed", False))
        
        ws.append(["Total Annotations:", total])
        ws.append(["Malformed:", malformed])
        ws.append(["Completed:", total - malformed])
        ws.append(["Malformed %:", f"{(malformed/total*100):.2f}%" if total > 0 else "0%"])
        ws.append([])
        
        # By domain
        ws.append(["By Domain:"])
        domains = {}
        for item in items:
            domain = item.get("domain", "Unknown")
            if domain not in domains:
                domains[domain] = {"total": 0, "malformed": 0}
            domains[domain]["total"] += 1
            if item.get("malformed", False):
                domains[domain]["malformed"] += 1
        
        ws.append(["Domain", "Total", "Malformed", "Completed"])
        for domain, stats in sorted(domains.items()):
            ws.append([domain, stats["total"], stats["malformed"], stats["total"] - stats["malformed"]])
        
        ws.append([])
        
        # By annotator
        ws.append(["By Annotator:"])
        annotators = {}
        for item in items:
            ann_id = item.get("annotator_id", 0)
            if ann_id not in annotators:
                annotators[ann_id] = {"total": 0, "malformed": 0}
            annotators[ann_id]["total"] += 1
            if item.get("malformed", False):
                annotators[ann_id]["malformed"] += 1
        
        ws.append(["Annotator", "Total", "Malformed", "Completed"])
        for ann_id, stats in sorted(annotators.items()):
            ws.append([f"Annotator {ann_id}", stats["total"], stats["malformed"], stats["total"] - stats["malformed"]])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
